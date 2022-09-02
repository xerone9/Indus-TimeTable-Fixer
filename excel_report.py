import openpyxl as xl
import os.path
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension, DimensionHolder
from openpyxl.utils import get_column_letter

list = ['09:00 - 10:00 ----- Interior Design Studio-I (Thy) ----- Sadia Perveen (Visiting) ----- Fashion theis room', '09:00 - 11:00 ----- Design Collection-I (Lab) ----- Faryal Ahsun (Fulltime) ----- Fashion thesis room', '09:00 - 11:00 ----- History of Art &amp; Culture-I ----- Zainab Abid (Visiting) ----- LEcture Room,LEcture Room', '09:00 - 11:00 ----- History of Art &amp; Culture-II ----- Afsheen Khalid (Visiting) ----- Lecture Room', '09:00 - 11:00 ----- Intro to Textile ----- Fareesa Javaid (Fulltime) ----- Lecture Room,Computer Lab', '09:00 - 12:00 ----- Business English,Business English ----- Yamna Khan (Visiting) ----- Lecture Room,LEcture Room', '09:00 - 12:00 ----- Digital Communication-II (Lab) ----- Amna Hashmi (Visiting) ----- Lecture Room,Lecture Room', '10:00 - 12:00 ----- Interior Design Studio-I (Lab) ----- Sadia Perveen (Visiting) ----- Seminar Hall', '10:00 - 13:00 ----- Collection-II (Thesis) (Lab) ----- Noor Ul Ain Shaikh (Visiting) ----- Textile Thesis Room', '11:00 - 12:00 ----- Design Collection-II (Lab) ----- Faryal Ahsun (Fulltime) ----- Textile thesis Lab', '12:30 - 14:30 ----- History of Art and Architecture-I ----- Kainat Riaz (Fulltime) ----- Thesis room,Lecture Room', '12:30 - 14:30 ----- History of Arts ----- Syeda Mona Batool Taqvi (Visiting) ----- lec room', '12:30 - 15:30 ----- Design Collection-III (Lab) ----- Faryal Ahsun (Fulltime) ----- Fashion theis room', '12:30 - 15:30 ----- English-I (Compulsory) ----- Zuraiz Akhter (Fulltime) ----- Fashion thesis room', '12:30 - 15:30 ----- History of Media Art ----- Syeda Mona Batool Taqvi (Visiting) ----- LEcture Room,LEcture Room', '12:30 - 15:30 ----- Media Laws and Ethics ----- Ayesha Naveed (Fulltime) ----- Lec room']

def time_in_range(start, end, current):
    return start <= current <= end


def set_border(ws, rows):
    top = 'A1:D1'
    bottom = 'A' + str(rows) + ':D' + str(rows)
    left = 'A1:A' + str(rows)
    right = 'D1:D' + str(rows)
    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    fix_1 = Border(top=thin, left=thin, right=thick, bottom=thick)
    fix_2 = Border(top=thin, left=thick, right=thin, bottom=thick)
    fix_3 = Border(top=thick, left=thick, right=thin, bottom=thin)
    fix_4 = Border(top=thick, left=thin, right=thick, bottom=thin)
    for row in ws[top]:
        for cell in row:
            cell.border = Border(top=thick, left=thin, right=thin, bottom=thin)
    for row in ws[bottom]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thick)
    for row in ws[left]:
        for cell in row:
            cell.border = Border(top=thin, left=thick, right=thin, bottom=thin)
    for row in ws[right]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thick, bottom=thin)
    ws.cell(row=int(rows), column=4).border = fix_1
    ws.cell(row=int(rows), column=1).border = fix_2
    ws.cell(row=1, column=1).border = fix_3
    ws.cell(row=1, column=4).border = fix_4

            # cell.border = Border(top=thick, left=thick, right=thick, bottom=thick)


def excel(list, day, session):
    desktop = os.path.expanduser("~\Desktop\\")
    saveLocation = desktop + "Scehedule For Day " + str(day) + ".xlsx"
    wb = xl.Workbook()
    wb.save(saveLocation)
    wb.create_sheet(str(day))
    sheet = wb[str(day)]
    ws = wb[str(day)]


    thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

    center = Alignment(horizontal='center', vertical='center')
    titleColor = PatternFill(start_color='000000',
                          end_color='000000',
                          fill_type='solid')
    headingColor = PatternFill(start_color='000000',
                          end_color='000000',
                          fill_type='solid')
    subHeadingColor = PatternFill(start_color='D0CECE',
                          end_color='D0CECE',
                          fill_type='solid')
    locationColor = PatternFill(start_color='00FFFF',
                          end_color='00FFFF',
                          fill_type='solid')
    deletionColor = PatternFill(start_color='ff0000',
                                end_color='ff0000',
                                fill_type='solid')

    titleFont = Font(color="FFFFFF", size=11)
    headerFont = Font(color="FFFFFF", size=11, bold=True)
    subHeadingFont = Font(color="000000", size=11, bold=True)
    headingFont = Font(color="FFFFFF", size=12, bold=True)

    cell = sheet.cell(1, 1)
    cell.value = day + "  Classes Time Table " + str(session)
    sheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=4)
    cell.fill = titleColor
    cell.font = headingFont
    cell.alignment = center
    cell = sheet.cell(4, 1)
    cell.value = "ClassTimings"
    # cell.font = Font(bold=True)
    cell.font = headerFont
    cell.fill = headingColor
    cell.border = thin_border
    cell.alignment = center
    cell = sheet.cell(4, 2)
    cell.value = "Course Title "
    cell.font = Font(bold=True)
    cell.fill = headingColor
    cell.font = headerFont
    cell.border = thin_border
    cell.alignment = center
    cell = sheet.cell(4, 3)
    cell.value = "Teacher"
    cell.font = headerFont
    cell.fill = headingColor
    cell.border = thin_border
    cell.alignment = center
    cell = sheet.cell(4, 4)
    cell.value = "Location"
    cell.font = headerFont
    cell.fill = headingColor
    cell.border = thin_border
    cell.alignment = center
    cell = sheet.cell(5, 1)
    cell.value = "First Slot"
    sheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=4)
    cell.fill = subHeadingColor
    cell.font = subHeadingFont
    cell.alignment = center
    wb.save(saveLocation)

    row = 6
    for entry in list:
        listed = entry.split(" ----- ")
        time = sheet.cell(row, 1)
        time.value = listed[0]
        time.border = thin_border
        time.alignment = center
        subject = sheet.cell(row, 2)
        subject.value = listed[1]
        subject.border = thin_border
        subject.alignment = center
        teacher = sheet.cell(row, 3)
        teacher.value = listed[2]
        teacher.border = thin_border
        teacher.alignment = center
        room = sheet.cell(row, 4)
        room.value = listed[3]
        room.border = thin_border
        room.alignment = center
        row += 1
    wb.save(saveLocation)

    number_12 = 0
    for row in range(6, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        splitem = str(cell.value).split(" - ")[0].split(":")
        start_time = (9, 0, 0)
        end_time = (11, 59, 0)
        current_time = (int(splitem[0]), int(splitem[1]), 0)
        check = time_in_range(start_time, end_time, current_time)

        if check is False:
            number_12 = row
            break

    if number_12 != 0:
        sheet.insert_rows(number_12)
        cell = sheet.cell(number_12, 1)
        cell.value = "Second Slot"
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell.fill = subHeadingColor
        cell.font = subHeadingFont
        cell.alignment = center

    number_15 = 0
    for row in range(6, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        if cell.value == "Second Slot":
            pass
        else:
            splitem = str(cell.value).split(" - ")[0].split(":")
            start_time = (9, 0, 0)
            end_time = (15, 29, 0)
            current_time = (int(splitem[0]), int(splitem[1]), 0)
            check = time_in_range(start_time, end_time, current_time)
            if check is False:
                number_15 = row
                break

    if number_15 != 0:
        sheet.insert_rows(number_15)
        cell = sheet.cell(number_15, 1)
        cell.value = "Third Slot"
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell.fill = subHeadingColor
        cell.font = subHeadingFont
        cell.alignment = center

    wb.save(saveLocation)

    subject_length = 0
    teacher_name_length = 0
    room_length = 0
    for row in range(1, sheet.max_row + 1):
        subject = sheet.cell(row, 2)
        teacher = sheet.cell(row, 3)
        room = sheet.cell(row, 4)
        try:
            if len(subject.value) > subject_length:
                subject_length = len(subject.value)
            if len(teacher.value) > teacher_name_length:
                teacher_name_length = len(teacher.value)
            if len(room.value) > room_length:
                room_length = len(room.value)
        except TypeError:
            pass

    wb.save(saveLocation)

    dim_holder = DimensionHolder(worksheet=ws)

    dim_holder[get_column_letter(1)] = ColumnDimension(ws, min=1, max=1, width=20)
    dim_holder[get_column_letter(2)] = ColumnDimension(ws, min=2, max=2, width=subject_length + 3)
    dim_holder[get_column_letter(3)] = ColumnDimension(ws, min=3, max=3, width=teacher_name_length + 3)
    dim_holder[get_column_letter(4)] = ColumnDimension(ws, min=4, max=4, width=room_length + 3)

    for row in range(4, sheet.max_row + 1):
        ws.row_dimensions[row].height = int(30)

    ws.column_dimensions = dim_holder
    wb.save(saveLocation)

    rows = 0
    for row in range(1, sheet.max_row + 1):
        rows += 1

    set_border(sheet, rows)
    del wb['Sheet']
    wb.save(saveLocation)





