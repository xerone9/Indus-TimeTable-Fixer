import openpyxl as xl
from tkinter import *
from tkinter import filedialog
import tkinter.font as tkFont
from tkinter import ttk
import datetime
import webbrowser
import os
from excel_report import excel
from pdf_report import generate_pdf_report
from datetime import date
import cryptocode


def callback(url):
    webbrowser.open_new(url)


def get_excel_report():
    for_the_day.config(fg='black')
    for_the_session.config(fg='black')
    if len(str(day.get())) < 1:
        for_the_day.config(fg='red')
    elif len(str(session.get())) < 1:
        for_the_session.config(fg='red')
    else:
        excel(final_table, str(day.get()), str(session.get()))
        desktop = os.path.expanduser("~\Desktop\\")
        savedLocation = desktop + "Scehedule For Day " + str(day.get()) + ".xlsx"
        os.startfile(savedLocation)



def get_pdf_report():
    for_the_day.config(fg='black')
    for_the_session.config(fg='black')
    if len(str(day.get())) < 1:
        for_the_day.config(fg='red')
    elif len(str(session.get())) < 1:
        for_the_session.config(fg='red')
    else:
        generate_pdf_report(final_table, str(day.get()), str(session.get()))
        desktop = os.path.expanduser("~\Desktop\\")
        savedLocation = desktop + "Scehedule For Day " + str(day.get()) + ".pdf"
        os.startfile(savedLocation)



def time_in_range(start, end, current):
    return start <= current <= end


def extra_time_in_range(start, current_start, current_end, end):
    return start <= current_start <= current_end <= end


def update_option():
    global teacher_selection
    variable.set(teacher_list[0])
    helv36 = tkFont.Font(family='Helvetica', size=15)
    teacher_selection = OptionMenu(root, variable, *teacher_list)
    teacher_selection.configure(cursor="hand1", font=helv36, bg="light blue", width=75)
    teacher_selection.place(x=20, y=180)

    helv20 = tkFont.Font(family='Helvetica', size=15)
    menu = root.nametowidget(teacher_selection.menuname)
    menu.config(font=helv20, bg="light blue")

    error.place_forget()


def add_entry():
    try:
        new_value = str(variable.get()).split(' ----- ')
        if str(str(new_value[0]) + " - " + variable2.get()) not in duplicate_check:
            status = False
            for i in duplicate_check:
                get_old_time = str(i).split(" -- ")
                if get_old_time[1] ==  variable2.get():
                    breakit = str(new_value[0]).split(" - ")
                    get_first_hour = str(breakit[0]).split(":")
                    current_hour = int(get_first_hour[0])
                    old_hour = str(get_old_time[0]).split(" - ")
                    get_old_start_hour = str(old_hour[0]).split(":")
                    old_start_hour = int(get_old_start_hour[0])
                    get_old_end_hour = str(old_hour[1]).split(":")
                    old_end_hour = int(get_old_end_hour[0])
                    start = datetime.time(old_start_hour, 0, 0)
                    end = datetime.time(old_end_hour, 0, 0)
                    current = datetime.time(current_hour, 1, 0)
                    status = time_in_range(start, end, current)
                    if status is True:
                        break

            if status is not True:
                inventory_stock.delete(*inventory_stock.get_children())
                values = str(variable.get() + ' ----- ' + variable2.get())
                final_table.append(values)
                final_table.sort()
                duplicate_check.append(str((new_value[0]) + " -- " + variable2.get()))


                count = 0
                for value in final_table:
                    entry = value.split(' ----- ')
                    inventory_stock.insert(parent="", index='end', iid=count, text=count + 1,
                                           values=(str(entry[0]), str(entry[1]), str(entry[2]), str(entry[3])))
                    count += 1

                teacher_list.remove(str(variable.get()))
                teacher_selection.destroy()
                update_option()
            else:
                error.config(text="Room busy")
                for children in inventory_stock.get_children():
                    child = inventory_stock.item(children)
                    for key, value in child.items():
                        if key == 'text':
                            count = int(value) - 1
                        if key == 'values':
                            if value[3] == variable2.get():
                                treeview_section = str(value[0]).split(" - ")[0].split(":")
                                treeview_start_hour = int(treeview_section[0])
                                treeview_start_min = int(treeview_section[1])
                                treeview_section = str(value[0]).split(" - ")[1].split(":")
                                treeview_end_hour = int(treeview_section[0])
                                treeview_end_min = int(treeview_section[1])
                                treeview_start = (treeview_start_hour, treeview_start_min, 0)
                                treeview_end = (treeview_end_hour, treeview_end_min, 0)
                                new_value_get = str(new_value[0]).split(" - ")[0].split(":")
                                new_value_get_end = str(new_value[0]).split(" - ")[1].split(":")
                                new_value_start_hour = int(new_value_get[0])
                                new_value_start_min = int(new_value_get[1])
                                new_value_start = (new_value_start_hour, new_value_start_min, 0)
                                status_start = time_in_range(treeview_start, treeview_end, new_value_start)
                                new_value_end_hour = int(new_value_get_end[0])
                                new_value_end_min = int(new_value_get_end[1])
                                new_value_end = (new_value_end_hour, new_value_end_min, 0)
                                status_end = time_in_range(treeview_start, treeview_end, new_value_end)
                                if status_start is True or status_end is True:
                                    child_id = inventory_stock.get_children()[count]
                                    inventory_stock.focus(child_id)
                                    inventory_stock.selection_set(child_id)

                error.place(x=600, y=230)


    except IndexError:
        frame.place(x=20, y=180)
        room_selection.place_forget()
        add_to_timetable.place_forget()
        excel_button.place(x=450, y=800)
        pdf_button.place(x=900, y=800)
        for_the_day.place(x=20, y=110)
        day.place(x=245, y=117)
        for_the_session.place(x=480, y=110)
        session.place(x=650, y=117)
        getTextLocation.place_forget()
        textLocation.place_forget()


def excel_working(location):
    browseButtonTXT.place_forget()
    wb = xl.load_workbook(location, data_only=True)
    sheet = wb.worksheets[0]

    # fetching common subjects
    subjects = []
    rooms = []

    with open('rooms.txt', 'r') as f:
        for line in f:
            rooms.append(line.rstrip())

    for row in range(2, sheet.max_row + 1):
        subject = sheet.cell(row, 2)
        if subject.value not in subjects:
            subjects.append(subject.value)


    # Fetching each subject time to analyse
    final_selection = []
    for item in subjects:
        start_time = "get value"
        for row in range(2, sheet.max_row + 1):
            subject = sheet.cell(row, 2)
            # Fixing Double Entries Because OF Break Up Time Slots
            if subject.value == item:
                time = sheet.cell(row, 1)
                if start_time == "get value":
                    start_time = str(time.value).split(" - ")[0]
                end_time = str(time.value).split(" - ")[1]
                correct_time = start_time + " - " + end_time
                teacher = sheet.cell(row, 3)
        final_selection.append(str(correct_time) + " ----- " + str(item) + " ----- " + str(teacher.value))
    return final_selection, rooms


def uploadExcel():
    filetypes = (
        ('Excel 2010 Files', '*.xlsx'),
    )

    filename = filedialog.askopenfilename(
        title='Open a file',
        filetypes=filetypes)
    textLocation.config(text=filename)

    for i in excel_working(filename)[0]:
        teacher_list.append(i)
    teacher_list.sort()

    for i in excel_working(filename)[1]:
        if i not in room_list:
            room_list.append(i)
    room_list.sort()

    global variable
    variable = StringVar(root)
    variable.set(teacher_list[0])
    var = IntVar()

    global variable2
    variable2 = StringVar(root)
    variable2.set(room_list[0])
    var = IntVar()

    helv36 = tkFont.Font(family='Helvetica', size=15)
    global teacher_selection
    teacher_selection = OptionMenu(root, variable, *teacher_list)
    teacher_selection.configure(cursor="hand1", font=helv36, bg="light blue", width=75)
    teacher_selection.place(x=20, y=180)

    helv20 = tkFont.Font(family='Helvetica', size=15)
    menu = root.nametowidget(teacher_selection.menuname)
    menu.config(font=helv20, bg="light blue")

    global room_selection
    room_selection = OptionMenu(root, variable2, *room_list)
    room_selection.configure(cursor="hand1", font=helv36, bg="light pink", width=30)
    room_selection.place(x=900, y=180)

    helv20 = tkFont.Font(family='Helvetica', size=15)
    menu = root.nametowidget(room_selection.menuname)
    menu.config(font=helv20, bg="light pink")

    global add_to_timetable
    add_to_timetable = Button(root, text='ADD', font=('Helvetica', 16, 'bold'), command=add_entry)
    add_to_timetable.config(fg='white', bg='black', width=10)
    add_to_timetable.place(x=1290, y=180)

    frame.place(x=20, y=300)



root = Tk()
root.resizable(0,0)
root.iconbitmap('icon.ico')
root.title('Class Scheduler Fixer - V-1.0')
root.geometry("1440x950+10+10")
root.configure(bg="white")

footer = Label(root, text="softwares.rubick.org", font=(14), cursor="hand2")
footer.bind("<Button-1>", lambda e: callback("https://github.com/xerone9/"))
footer.configure(foreground="white")
footer.configure(bg="black")
footer.pack(side=BOTTOM)

img=PhotoImage(file='iu_logo.png')
label = Label(root, image=img)
label.configure(foreground="black")
label.configure(bg="white")
label.place(x=10, y=2)

try:
    with open('license.ini', 'r') as f:
        for line in f:
            date_hash = str(line).replace("terminal = ", "")

    encoded = date_hash

    ## And then to decode it:
    decoded = cryptocode.decrypt(encoded, "rubick")
    split_date = str(decoded).split(" ")

    year = int(split_date[0])
    month = int(split_date[1])
    day = int(split_date[2])

    license_not_expire = date.today() < date(year, month, day)

    if license_not_expire is True:
        getTextLocation = Label(root, text="Select Excel File", font=("Comic Sans MS", 25, 'bold'))
        getTextLocation.configure(bg="white")
        getTextLocation.place(x=20, y=110)

        browseButtonTXT = Button(root, text='Select File', font=(20), command=uploadExcel)
        browseButtonTXT.place(x=350, y=120)

        textLocation = Label(root, text="", font=(10))
        textLocation.config(bg='white')
        textLocation.place(x=470, y=123)

        error = Label(root, text="", font=("Comic Sans MS", 35, 'bold'))
        error.config(bg='white', fg='red')
        error.place_forget()

        frame = LabelFrame(root, text="Current Time Table", padx=10, pady=10, bg="white", font=("Comic Sans MS", 25, 'bold'))
        frame.place_forget()

        style = ttk.Style()
        style.configure("Treeview.Heading", font=(None, 20))
        style.configure("Treeview", font=(None, 15), rowheight=30)
        style.map('Treeview', background=[('selected', '#FF0000')])

        inventory_stock = ttk.Treeview(frame, height=15, selectmode='none')
        inventory_stock['columns'] = ("Time Slot", "Teacher Name", "Subject", "Room")
        inventory_stock.column("#0", minwidth=25, width=100)
        inventory_stock.column("Time Slot", anchor=W, width=150)
        inventory_stock.column("Teacher Name", anchor=W, width=350)
        inventory_stock.column("Subject", anchor=W, width=350)
        inventory_stock.column("Room", anchor=CENTER, width=400)

        inventory_stock.heading("#0", text="S. No", anchor=W)
        inventory_stock.heading("Time Slot", text="Time Slot", anchor=W)
        inventory_stock.heading("Teacher Name", text="Teacher Name", anchor=W)
        inventory_stock.heading("Subject", text="Subject", anchor=W)
        inventory_stock.heading("Room", text="Room", anchor=CENTER)
        inventory_stock.tag_configure('T', font='Arial 20')

        inventory_stock.pack(side='left', fill='y')

        scrollbar = Scrollbar(frame, orient="vertical", command=inventory_stock.yview)
        scrollbar.pack(side="right", fill="y")
        inventory_stock.configure(yscrollcommand=scrollbar.set)

        excel_button = Button(root, text='Get Excel', font=('Helvetica', 16, 'bold'), command=get_excel_report)
        excel_button.config(fg='white', bg='Green', width=10)
        excel_button.place_forget()

        pdf_button = Button(root, text='Get PDF', font=('Helvetica', 16, 'bold'), command=get_pdf_report)
        pdf_button.config(fg='white', bg='red', width=10)
        pdf_button.place_forget()

        for_the_day = Label(root, text="For the Day: ", font=("Comic Sans MS", 25, 'bold'))
        for_the_day.configure(bg="white")
        for_the_day.place_forget()

        day = Entry(root, font=("Comic Sans MS", 25, 'bold'), width=9)
        day.configure(bg="light grey", fg='blue')
        day.place_forget()

        for_the_session = Label(root, text="Session: ", font=("Comic Sans MS", 25, 'bold'))
        for_the_session.configure(bg="white")
        for_the_session.place_forget()

        session = Entry(root, font=("Comic Sans MS", 25, 'bold'), width=20)
        session.configure(bg="light grey", fg='blue')
        session.place_forget()


        teacher_list = []
        room_list = []
        final_table = []
        duplicate_check = []

    else:
        error = Label(root, text="Software Expired :(", font=("Comic Sans MS", 35, 'bold'))
        error.config(bg='white', fg='red')
        error.place(relx=0.35, rely=0.4)
except FileNotFoundError:
    error = Label(root, text="Licensce File Missing :(", font=("Comic Sans MS", 35, 'bold'))
    error.config(bg='white', fg='red')
    error.place(relx=0.35, rely=0.4)
except ValueError:
    error = Label(root, text="Licensce Corrupted :(", font=("Comic Sans MS", 35, 'bold'))
    error.config(bg='white', fg='red')
    error.place(relx=0.35, rely=0.4)


root.mainloop()




